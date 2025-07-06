import os
import re
import time
import pandas as pd
from datetime import datetime
import numpy as np
from flask import Flask, render_template, request, jsonify, send_file
from werkzeug.utils import secure_filename
from flask_socketio import SocketIO
import os.path
import atexit
import shutil
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import psutil

global temp_output
application = Flask(__name__)
application.config['MAX_CONTENT_LENGTH'] = 1024 * 1024 * 1024  # 1GB

# Key changes for Posit Connect compatibility
socketio = SocketIO(
    application,
    async_mode='threading',  # Changed from 'eventlet' to 'threading'
    cors_allowed_origins="*",
    ping_timeout=120,
    ping_interval=60,
    max_http_buffer_size=1024 * 1024 * 1024,
    async_handlers=True,
    logger=True,
    engineio_logger=True
)

application.config['UPLOAD_FOLDER'] = 'uploads'
application.config['OUTPUT'] = 'outputs'
application.config['STATIC_FOLDER'] = 'static'

application.secret_key = 'supersecretkey'
application.config['TEMPLATES_AUTO_RELOAD'] = True
application.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0

current_process = None

@application.route('/health')
def health():
    return jsonify({"status": "healthy"}), 200

def cleanup_folders(max_age=3600):
    current_time = time.time()
    folders = [application.config['UPLOAD_FOLDER'], application.config['OUTPUT']]
    
    for folder in folders:
        try:
            if os.path.exists(folder):
                # Clean up old files
                for filename in os.listdir(folder):
                    filepath = os.path.join(folder, filename)
                    file_age = current_time - os.path.getmtime(filepath)
                    if file_age > max_age:
                            if os.path.isfile(filepath):
                                os.unlink(filepath)
                            elif os.path.isdir(filepath):
                               shutil.rmtree(filepath)
            else:
                os.makedirs(folder)
        except Exception as e:
            print(f"Error cleaning up {folder}: {str(e)}")

# Register cleanup for application shutdown
atexit.register(cleanup_folders)

# Clean up on startup
cleanup_folders()

@application.before_request
def ensure_folders_exist():
    """Ensure upload and output folders exist"""
    for folder in [application.config['UPLOAD_FOLDER'], application.config['OUTPUT']]:
        if not os.path.exists(folder):
            os.makedirs(folder)


def send_progress_update(message, category='info', tab=None):
    if not hasattr(send_progress_update, 'message_counter'):
        send_progress_update.message_counter = 0
    send_progress_update.message_counter += 1
    
    # Add timestamp
    timestamp = datetime.now().strftime('%H:%M:%S')
    
    # Format message based on category
    colors = {
        'error': 'red',
        'success': 'green',
        'warning': 'orange',
        'original': 'blue',
        'duplicate': 'blue',
        'info': 'black',
        'time': '#666666'  # Gray color for timestamps
    }
    
    color = colors.get(category, 'black')
    formatted_message = f'<span style="color: {colors["time"]}">[{timestamp}]</span> <span style="color: {color};">{message}</span>'

    # Emit message with ID and tab information
    socketio.emit('progress_update', {
        'message': formatted_message,
        'category': category,
        'message_id': send_progress_update.message_counter,
        'tab': tab  # Add tab information
    })
    # Remove socketio.sleep(0) - not needed with threading mode
    # socketio.sleep(0)


def save_uploaded_file(file):
    filename = secure_filename(file.filename)
    filepath = os.path.join(application.config['UPLOAD_FOLDER'], filename)
    file.save(filepath)
    return filepath


def find_last_column(sheet):
    return sheet.max_column


def check_source_file(sheet, sheet_type):
    errors = []
    # Reading cell values directly from openpyxl
    ep_type_in_line3 = sheet['B3'].value
    if ep_type_in_line3 is None:
        errors.append(f"Sheet {sheet.title}: Line 3 is empty or invalid.")
        return errors

    if sheet_type == 'OEP' and 'OEP' not in str(ep_type_in_line3):
        errors.append(f"Sheet {sheet.title}: Expected 'OEP' in line 3, found '{ep_type_in_line3}'")
    elif sheet_type == 'AEP' and 'AEP' not in str(ep_type_in_line3):
        errors.append(f"Sheet {sheet.title}: Expected 'AEP' in line 3, found '{ep_type_in_line3}'")

    try:
        peril_in_line3 = str(ep_type_in_line3).split(' - ')[2].strip()
        peril_map = {'FL': 'Flood'}
        expected_peril = peril_map.get(peril_in_line3, peril_in_line3)
    except (AttributeError, IndexError, TypeError):
        errors.append(f"Sheet {sheet.title}: Unable to extract peril from line 3")
        return errors

    last_col = find_last_column(sheet)
    # perils in row 6 from col 3 to last_col
    perils_in_row_6 = []
    for col in range(3, last_col+1):
        val = sheet.cell(row=6, column=col).value
        if val is not None:
            perils_in_row_6.append(val.strip())

    peril_map = {'FL': 'Flood'}
    perils_in_row_6_mapped = [peril_map.get(p, p) for p in perils_in_row_6]

    if not all(p == expected_peril for p in perils_in_row_6_mapped):
        errors.append(f"Sheet {sheet.title}: Peril mismatch in row 6 - Expected '{expected_peril}', found {set(perils_in_row_6)}")

    # original/adjusted checks (row 23 and row 30)
    original_adjusted_23 = [sheet.cell(row=23, column=col).value for col in range(3, last_col+1)]
    original_adjusted_30 = [sheet.cell(row=30, column=col).value for col in range(3, last_col+1)]

    for i, val in enumerate(original_adjusted_23):
        if val is not None and original_adjusted_30[i] is not None:
            val_lower = str(val).lower()
            row30_val_lower = str(original_adjusted_30[i]).lower()
            if val_lower == 'original' and 'original' not in row30_val_lower:
                errors.append(f"Sheet {sheet.title}: Mismatch at column {i+3} - Expected 'original' in row 30, found '{original_adjusted_30[i]}'")
            elif val_lower == 'adjusted' and 'adjusted' not in row30_val_lower:
                errors.append(f"Sheet {sheet.title}: Mismatch at column {i+3} - Expected 'adjusted' in row 30, found '{original_adjusted_30[i]}'")

    modelling_id_24 = sheet['D24'].value
    modelling_id_29 = sheet['D29'].value
    if modelling_id_24 != modelling_id_29:
        errors.append(f"Sheet {sheet.title}: Modelling ID mismatch - Line 24: {modelling_id_24}, Line 29: {modelling_id_29}")

    years_in_row_9 = [sheet.cell(row=9, column=c).value for c in range(2, last_col+1)]
    years_in_row_31 = [sheet.cell(row=31, column=c).value for c in range(2, last_col+1)]

    for i, (year_9, year_31) in enumerate(zip(years_in_row_9, years_in_row_31)):
        if year_9 and year_31:
            if isinstance(year_9, datetime):
                year_9_extracted = year_9.year
            else:
                try:
                    year_9_extracted = int(str(year_9).split('/')[-1].strip())
                except (ValueError, IndexError, AttributeError):
                    errors.append(f"Sheet {sheet.title}: Invalid year format in row 9, column {i+2}")
                    continue
            year_31_str = str(int(year_31)) if isinstance(year_31, float) else str(year_31).strip()
            if str(year_9_extracted) != year_31_str:
                errors.append(f"Sheet {sheet.title}: Year mismatch at column {i+2} - Row 9: {year_9_extracted}, Row 31: {year_31}")

    return errors


def contains_year(df, row, col, year_to_find):
            cell = df.iloc[row, col]
            if isinstance(cell, datetime) and cell.year == year_to_find:
                if col + 1 < len(df.columns):
                    right_cell = df.iloc[row, col + 1]
                    if isinstance(right_cell, datetime) and right_cell.year == cell.year:
                        if row + 14 < len(df):
                            below_cell1 = df.iloc[row + 14, col]
                            below_right_cell1 = df.iloc[row + 14, col + 1]
                            if below_cell1 == below_right_cell1:
                                if row + 15 < len(df):
                                    below_cell2 = df.iloc[row + 15, col]
                                    below_right_cell2 = df.iloc[row + 15, col + 1]
                                    if below_cell2 == below_right_cell2:
                                        return False
                                else:
                                    below_cell2 = re.sub(r'[^\d\+\-\*/]', '', str(df.iloc[row + 15, col]))
                                    below_right_cell2 = re.sub(r'[^\d\+\-\*/]', '', str(df.iloc[row + 15, col + 1]))
                                    if below_cell2 == below_right_cell2:
                                        return False
                return True
            return False


def post_process_duplicates(file_path):
    df = pd.read_excel(file_path, sheet_name="Processed Data", engine='openpyxl')
    
    # Define the columns to check for duplicates
    duplicate_columns = [0, 1, 3, 4, 5, 9, 10, 11, 16, 17]
    duplicates = df[df.duplicated(subset=df.columns[duplicate_columns], keep=False)]
    column_names = df.columns[duplicate_columns].tolist()
    
    send_progress_update("Checking for duplicates...", 'info', 'postprocess')
    send_progress_update(f"Columns being checked: {column_names}", 'info', 'postprocess')
    
    if not duplicates.empty:
        send_progress_update("Duplicates found:", "warning", 'postprocess')
        
        # Group duplicates
        grouped = duplicates.groupby(list(df.columns[duplicate_columns]))
        
        for name, group in grouped:
            original_row_index = group.index[0]
            send_progress_update(f"Original Row Number: {original_row_index + 2}", "info", 'postprocess')
            for row_index, row in group.iloc[1:].iterrows():
                send_progress_update(
                    f"Duplicate Row Number: {row_index + 2} -> "
                    f"Peril: {row['Peril']}, "
                    f"Portfolio: {row['Portfolio']}, "
                    f"Original/Adjusted: {row['original/adjusted']}, "
                    f"Measure: {row['Measure']}, "
                    f"NatCat Model: {row['NatCat Model']}, "
                    f"Return Period: {row['Return Period']}, "
                    f"Value: {row['Value']}",
                    "duplicate",
                    'postprocess'
                )
    else:
        send_progress_update("No duplicates found", "success", 'postprocess')


def process_data_48(dataframe, year_to_find, setup, column_mapping, columns_in_table):
    values_dict_column_b = {}
    values_dict_column_d = {}
    column_b = dataframe.iloc[:, 1].astype(str)
    column_d = dataframe.iloc[:, 3].astype(str)
    year_indices = []
    row_number = 26

    for column_index, column in enumerate(dataframe.columns):
        for row_index in range(len(dataframe)):
            if contains_year(dataframe, row_index, column_index, year_to_find):
                year_indices.append((column_index - 1) if setup =='OEP' else (column_index)) # check why this is the case?
                for index, value in enumerate(column_b):
                    if not pd.isna(value) and 3 <= index <= 25:
                        variable_name = re.sub(r'\W+', '_', value)
                        values_dict_column_b.setdefault(variable_name, []).append(dataframe.iloc[index, column_index])
                for index, value in enumerate(column_d):
                    if not pd.isna(value) and 31 <= index <= 45:
                        variable_name = re.sub(r'\W+', '_', value)
                        values_dict_column_d.setdefault(variable_name, []).append(dataframe.iloc[index, column_index])
                break

    if not year_indices:
        socketio.emit('progress_update', {'message': f"No matching year found in sheet for year {year_to_find}", 'category': 'error'})
        print (f"No matching year found in sheet for year {year_to_find}")
        # Remove socketio.sleep(0)
        return pd.DataFrame()

    dataframe.iloc[[row_number], year_indices] = dataframe.iloc[[row_number], year_indices].infer_objects(copy=False)
    Portfolio = dataframe.iloc[[row_number], year_indices].ffill(axis=1).values.flatten()  # fills an empty value with the last value
    df_portfolio = pd.DataFrame({'Portfolio': Portfolio})
    print(Portfolio)
    df_portfolio.to_csv('portfolio.csv', index=False)
    result_dataframe = pd.DataFrame.from_dict(values_dict_column_b, orient='index').T
    result_dataframe['Portfolio'] = Portfolio

    if setup == 'AEP':
        desired_order = ['1000', '500', '250', '200', '100', '50', '25', '10', '5']
    else:
        desired_order = ['1000', '500', '250', '200', '100', '50', '25', '10', '5', 'Exposure', 'Modelled_Exposure', 'Average_Annual_Loss']


    column_d_reordered = {key: values_dict_column_d[key] for key in desired_order}
    repeat_count = len(column_d_reordered)
    repeated_data = [result_dataframe.loc[[index]].reindex([index] * repeat_count) for index in result_dataframe.index]
    repeated_dataframe = pd.concat(repeated_data, ignore_index=True)

    Return_Period = list(column_d_reordered.keys()) * len(result_dataframe)
    Value = np.array(list(column_d_reordered.values())).T.ravel()

    if len(Return_Period) != len(Value):
        socketio.emit('progress_update', {'message': "Error: Length of 'Return Period' does not match 'Value'", 'category': 'error'})
        print( "Error: Length of 'Return Period' does not match 'Value'")
        # Remove socketio.sleep(0)
        return pd.DataFrame()
    if len(Return_Period) != len(column_d_reordered) * len(result_dataframe):
        socketio.emit('progress_update', {'message': "Error: Length of variables does not match the number of lines of businesses", 'category': 'error'})
        print("Error: Length of variables does not match the number of lines of businesses")
        # Remove socketio.sleep(0)
        return pd.DataFrame()
    if len(repeated_dataframe) != len(Value):
        socketio.emit('progress_update', {'message': "Error: Length of repeated data and return periods don't match", 'category': 'error'})
        print ("Error: Length of repeated data and return periods don't match")
        # Remove socketio.sleep(0)
        return pd.DataFrame()

    dataframe_final = repeated_dataframe.copy()
    dataframe_final['Return Period'] = Return_Period
    dataframe_final['Value'] = Value
    dataframe_final_renamed = dataframe_final.rename(columns=column_mapping)


    dataframe_final_renamed2 = dataframe_final_renamed[columns_in_table]
    return dataframe_final_renamed2

def process_data_50(dataframe, year_to_find, setup, column_mapping, columns_in_table):
    values_dict_column_b = {}
    values_dict_column_d = {}
    column_b = dataframe.iloc[:, 1].astype(str)
    column_d = dataframe.iloc[:, 3].astype(str)
    year_indices = []
    row_number = 26

    for column_index, column in enumerate(dataframe.columns):
        for row_index in range(len(dataframe)):
            if contains_year(dataframe, row_index, column_index, year_to_find):
                year_indices.append((column_index - 1) if setup =='OEP' else (column_index)) # check why this is the case?
                for index, value in enumerate(column_b):
                    if not pd.isna(value) and 3 <= index <= 25:
                        variable_name = re.sub(r'\W+', '_', value)
                        values_dict_column_b.setdefault(variable_name, []).append(dataframe.iloc[index, column_index])
                for index, value in enumerate(column_d):
                    if not pd.isna(value) and 31 <= index <= 47:
                        variable_name = re.sub(r'\W+', '_', value)
                        values_dict_column_d.setdefault(variable_name, []).append(dataframe.iloc[index, column_index])
                break

    if not year_indices:
        socketio.emit('progress_update', {'message': f"No matching year found in sheet for year {year_to_find}", 'category': 'error'})
        print (f"No matching year found in sheet for year {year_to_find}")
        # Remove socketio.sleep(0)
        return pd.DataFrame()

    dataframe.iloc[[row_number], year_indices] = dataframe.iloc[[row_number], year_indices].infer_objects(copy=False)    
    Portfolio = dataframe.iloc[[row_number], year_indices].ffill(axis=1).values.flatten()  # fills an empty value with the last value
    
    send_progress_update(f"Number of years: {len(year_indices)}", 'info')

    result_dataframe = pd.DataFrame.from_dict(values_dict_column_b, orient='index').T
    result_dataframe['Portfolio'] = Portfolio
    
    df_portfolio = pd.DataFrame({'Portfolio': Portfolio})
    df_portfolio.to_csv('portfolio.csv', index=False)

    if setup == 'AEP':
        desired_order = ['1000', '500', '250', '200', '100', '50', '25', '10', '5']
    else:
        desired_order = [
            '1000', '500', '250', '200', '100', '50', '25', '10', '5',
            'Exposure', 'Modelled_Exposure', 'Average_Annual_Loss',
            'Standard_Deviation', 'Coefficient_of_Variation'
        ]

    column_d_reordered = {key: values_dict_column_d[key] for key in desired_order}
    repeat_count = len(column_d_reordered)
    repeated_data = [result_dataframe.loc[[index]].reindex([index] * repeat_count) for index in result_dataframe.index]
    repeated_dataframe = pd.concat(repeated_data, ignore_index=True)

    Return_Period = list(column_d_reordered.keys()) * len(result_dataframe)
    Value = np.array(list(column_d_reordered.values())).T.ravel()
    if len(Return_Period) != len(Value):
        socketio.emit('progress_update', {'message': "Error: Length of 'Return Period' does not match 'Value'", 'category': 'error'})
        print( "Error: Length of 'Return Period' does not match 'Value'")
        # Remove socketio.sleep(0)
        return pd.DataFrame()
    if len(Return_Period) != len(column_d_reordered) * len(result_dataframe):
        socketio.emit('progress_update', {'message': "Error: Length of variables does not match the number of lines of businesses", 'category': 'error'})
        print("Error: Length of variables does not match the number of lines of businesses")
        # Remove socketio.sleep(0)
        return pd.DataFrame()
    if len(repeated_dataframe) != len(Value):
        socketio.emit('progress_update', {'message': "Error: Length of repeated data and return periods don't match", 'category': 'info'})
        print ("Error: Length of repeated data and return periods don't match")
        # Remove socketio.sleep(0)
        return pd.DataFrame()

    dataframe_final = repeated_dataframe.copy()
    dataframe_final['Return Period'] = Return_Period
    dataframe_final['Value'] = Value
    dataframe_final_renamed = dataframe_final.rename(columns=column_mapping)


    dataframe_final_renamed2 = dataframe_final_renamed[columns_in_table]
    return dataframe_final_renamed2


def process_data_65(dataframe, year_to_find, setup, column_mapping, columns_in_table):
    processed_data1 = process_data_48(dataframe, year_to_find, setup, column_mapping, columns_in_table)
    df2 = dataframe.copy()
    df2.iloc[35:46] = dataframe.iloc[55:66]
    processed_data2 = process_data_48(df2, year_to_find, setup, column_mapping, columns_in_table)
    if not processed_data2.empty:
        processed_data2['Perspective'] = 'Net Pre CAT'
    concatenated_data = pd.concat([processed_data1, processed_data2], ignore_index=True)
    return concatenated_data


def copy_worksheet(source_sheet, target_wb, title):
    # Create new sheet in target workbook
    new_sheet = target_wb.create_sheet(title=title)
    # Copy cell values
    for row in source_sheet.iter_rows():
        for cell in row:
            new_sheet.cell(row=cell.row, column=cell.column, value=cell.value)


@application.route('/')
def index():
    return render_template('index.html')


@application.route('/process', methods=['POST'])
def process():
    try:
        source_file = request.files.get('source')
        year = request.form.get('year')

        if not source_file:
            send_progress_update("No source file uploaded.", 'error', 'convert')
            return jsonify({'status': 'error', 'message': 'No source file uploaded.'})
        
        if not year or not year.isdigit():
            send_progress_update("Invalid or missing year.", 'error', 'convert')
            return jsonify({'status': 'error', 'message': 'Invalid or missing year.'})

        year = int(year)
        temp_source = save_uploaded_file(source_file)
        
        send_progress_update("File uploaded successfully.", 'info', 'convert')

        # Load source workbook
        try:
            wb1 = load_workbook(temp_source, data_only=True)
            send_progress_update("Workbook loaded successfully.", 'info', 'convert')
        except Exception as e:
            send_progress_update(f"Error loading workbook: {str(e)}", 'error', 'convert')
            return jsonify({'status': 'error', 'message': f'Error loading workbook: {str(e)}'})

        # Get sheets to process
        sheets_to_process = [s for s in wb1.sheetnames 
                           if wb1[s].sheet_state == 'visible' 
                           and (s.rstrip().endswith("AEP") or s.rstrip().endswith("OEP"))]
        
        send_progress_update(f"Found {len(sheets_to_process)} sheets to process.", 'info')

        # Initialize list for processed data
        all_processed_data = []
        column_mapping = {
            'Business_Unit_BU_': 'Business Unit',
            'incl_Subperil': 'incl Subperil',
            'Country_modelled_': 'Country modelled',
            'Date_of_Portfolio': 'Date of Portfolio',
            'Measure_Perspective': 'Perspective',
            'Exchange_Rate': 'Exchange Rate',
            'Data_Supplier': 'Data Supplier',
            'NatCat_Model': 'NatCat Model',
            'Model_Version': 'Model Version',
            'Post_loss_amplification': 'Post Loss Amplification',
            'Original_adjusted': 'original/adjusted'
        }
        columns_in_table = [
            'Business Unit', 'Peril', 'incl Subperil', 'Portfolio', 'original/adjusted',
            'Modelling_ID', 'Country modelled', 'Date of Portfolio', 'Perspective',
            'Measure', 'Return Period', 'Value', 'Currency', 'Exchange Rate', 'Data Supplier',
            'Modeler', 'NatCat Model', 'Model Version', 'Post Loss Amplification', 'Comments'
        ]

        all_processed_data = []

        # Process each sheet
        for sheet_name in sheets_to_process:
            try:
                send_progress_update(f"Processing sheet: {sheet_name}", 'info')
                
                # Read sheet into DataFrame
                df_sheet = pd.read_excel(temp_source, sheet_name=sheet_name, engine='openpyxl')
                # Find AAL row
                search_word = "AAL"
                aal_rows = df_sheet[df_sheet.apply(lambda row: row.astype(str).str.contains(search_word, case=False, na=False).any(), axis=1)].index
                row_number = aal_rows.max() + 1

                # Process data based on row number
                sheet_type = 'AEP' if sheet_name.rstrip().endswith("AEP") else 'OEP'
                processed_data = None
                
                if row_number == 48:
                    processed_data = process_data_48(df_sheet, year, sheet_type, column_mapping, columns_in_table)
                elif row_number == 50:
                    processed_data = process_data_50(df_sheet, year, sheet_type, column_mapping, columns_in_table)
                elif row_number == 68:
                    processed_data = process_data_65(df_sheet, year, sheet_type, column_mapping, columns_in_table)
                else:
                    send_progress_update(f"Unsupported sheet format (row_number={row_number})", 'warning')
                    continue

                if processed_data is not None and not processed_data.empty:
                    all_processed_data.append(processed_data)
                    send_progress_update(f"Successfully processed sheet: {sheet_name}", 'success')
                else:
                    send_progress_update(f"No data processed for sheet: {sheet_name}", 'warning')

            except Exception as e:
                send_progress_update(f"Error processing sheet {sheet_name}: {str(e)}", 'error')
                continue

        if all_processed_data:
            try:
                # Combine all processed data
                final_data = pd.concat(all_processed_data, ignore_index=True)
                send_progress_update(f"Combined all processed data. Total rows: {len(final_data)}.", 'info')

                # Prepare output file
                input_filename = os.path.splitext(os.path.basename(source_file.filename))[0]
                output_filename = f"{input_filename}_database_format.xlsx"
                output_path = os.path.join(application.config['OUTPUT'], output_filename)

                # Create Excel writer
                with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                    # Write processed data
                    send_progress_update("Writing processed data...", 'info')
                    final_data.to_excel(writer, sheet_name='Processed Data', index=False)
                    
                send_progress_update("File processing completed successfully.", 'success')
                
                return send_file(
                    output_path,
                    as_attachment=True,
                    download_name=output_filename,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )

            except Exception as e:
                send_progress_update(f"Error saving output file: {str(e)}", 'error')
                return jsonify({'status': 'error', 'message': str(e)})
        else:
            send_progress_update("No data was processed successfully.", 'error')
            return jsonify({'status': 'error', 'message': 'No data processed'})

    except Exception as e:
        send_progress_update(f"Unexpected error: {str(e)}", 'error')
        return jsonify({'status': 'error', 'message': str(e)})


@application.route('/preprocess', methods=['POST'])
def preprocess():
    try:
        if 'file' not in request.files:
            return jsonify({'status': 'error', 'message': 'No file uploaded'})
        
        file = request.files['file']
        if not file.filename:
            return jsonify({'status': 'error', 'message': 'No file selected'})        
            
        filepath = save_uploaded_file(file)
        
        # Add error handling for file loading
        try:
            wb = load_workbook(filepath, data_only=True)
        except Exception as e:
            send_progress_update(f"Error loading workbook: {str(e)}", 'error', 'preprocess')
            return jsonify({'status': 'error', 'message': f'Error loading workbook: {str(e)}'})
            
        all_errors = []
        
        sheets_to_process = [s for s in wb.sheetnames 
                           if wb[s].sheet_state == 'visible' 
                           and (s.rstrip().endswith("AEP") or s.rstrip().endswith("OEP"))]
        
        if not sheets_to_process:
            send_progress_update("No valid sheets found to process.", 'warning', 'preprocess')
            return jsonify({'status': 'warning', 'message': 'No valid sheets found'})
            
        send_progress_update(f"Found {len(sheets_to_process)} sheets to process.", 'info', 'preprocess')
        
        for sheet_name in sheets_to_process:
            try:
                sheet_type = 'AEP' if sheet_name.rstrip().endswith("AEP") else 'OEP'
                send_progress_update(f"Checking sheet: {sheet_name}", 'info', 'preprocess')
                sheet = wb[sheet_name]
                errors = check_source_file(sheet, sheet_type)
                if errors:
                    for error in errors:
                        send_progress_update(error, 'error', 'preprocess')
                    all_errors.extend(errors)
            except Exception as e:
                send_progress_update(f"Error processing sheet {sheet_name}: {str(e)}", 'error', 'preprocess')
                all_errors.append(f"Sheet {sheet_name}: {str(e)}")
        
        wb.close()
        
        if not all_errors:
            send_progress_update("Pre-processing completed successfully.", 'success', 'preprocess')
        else:
            send_progress_update("Pre-processing completed with errors.", 'warning', 'preprocess')
            
        return jsonify({
            'status': 'success' if not all_errors else 'warning',
            'errors': all_errors
        })
        
    except Exception as e:
        send_progress_update(f"Error during pre-processing: {str(e)}", 'error', 'preprocess')
        return jsonify({'status': 'error', 'message': str(e)})


@application.route('/postprocess', methods=['POST'])
def postprocess():
    try:
        if 'file' not in request.files:
            return jsonify({'status': 'error', 'message': 'No file uploaded'})
        
        file = request.files['file']
        if not file.filename:
            return jsonify({'status': 'error', 'message': 'No file selected'})
                    
        filepath = save_uploaded_file(file)
        
        post_process_duplicates(filepath)
        
        send_progress_update("Post-processing completed successfully.", 'success', 'postprocess')
        return jsonify({'status': 'success'})
        
    except Exception as e:
        send_progress_update(f"Error during post-processing: {str(e)}", 'error', 'postprocess')
        return jsonify({'status': 'error', 'message': str(e)})


@application.after_request
def cleanup_after_request(response):
    """Clean up old files after each request"""
    cleanup_folders(max_age=3600)  # Clean files older than 1 hour
    return response


@application.route('/help')
def help():
    try:
        return send_file(
            'static/docs/user_guide.pdf',
            mimetype='application/pdf',
            as_attachment=False
        )
    except Exception as e:
        return jsonify({'status': 'error', 'message': 'Help document not found'}), 404

# Make sure static folders exist
@application.before_request
def create_static_folders():
    static_folder = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', 'docs')
    if not os.path.exists(static_folder):
        os.makedirs(static_folder)

@application.route('/clear-cache', methods=['POST'])
def clear_cache():
    try:
        # Clear uploads folder
        for folder in [application.config['UPLOAD_FOLDER'], application.config['OUTPUT']]:
            if os.path.exists(folder):
                for filename in os.listdir(folder):
                    filepath = os.path.join(folder, filename)
                    try:
                        if os.path.isfile(filepath):
                            os.unlink(filepath)
                        elif os.path.isdir(filepath):
                            shutil.rmtree(filepath)
                    except Exception as e:
                        print(f"Error deleting {filepath}: {str(e)}")
        
        return jsonify({'status': 'success', 'message': 'Cache cleared successfully'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)})

@application.route('/end-process', methods=['POST'])
def end_process():
    try:
        # Get current process
        current_pid = os.getpid()
        current_process = psutil.Process(current_pid)
        
        # Terminate all child processes
        children = current_process.children(recursive=True)
        for child in children:
            child.terminate()
        
        # Clear any ongoing operations
        socketio.emit('process_terminated', {
            'message': 'Process terminated by user',
            'category': 'warning'
        })
        
        return jsonify({'status': 'success', 'message': 'Process terminated'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)})
    
@application.route('/.well-known/pki-validation/09C22592867EA83F056E3A74B19BAF0E.txt')
def serve_txt():
    content = """D13977FC6651419966CB4E112E3225B3E10BD34BF490ADDD37E7427ECB384E8F
comodoca.com
1ae91d39912bc2c"""
    return application.response_class(content, mimetype='text/plain')

# Modified for Posit Connect compatibility
app = application

# Only run socketio when executing directly (not through Posit Connect)
if __name__ == '__main__':
    socketio.run(application, debug=True)